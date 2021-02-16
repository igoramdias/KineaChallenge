from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formula import Tokenizer
from datetime import datetime as dt
from openpyxl import load_workbook
from pandas import Index
from ast import literal_eval
import pandas
import numpy
import shutil
import time
import os

def get_path(): #Pega o caminho atual do .py
  global file_path

  file_path = os.getcwd()

def contains_ticker_ipca(ticker):
  global ativos_incentivada
  
  try:
    idx = ativos_incentivada.get_loc(ticker)
  
  except KeyError:
    return False

  return idx

def contains_ticker_di(ticker):
  global ativos_convencional
  
  try:
    idx = ativos_convencional.get_loc(ticker)
  
  except KeyError:
    return False

  return idx

def read_base_debentures():
  global base_incentivada
  global base_convencional
  global ativos_incentivada
  global ativos_convencional
  
  base_incentivada = pandas.read_excel("P:\\PERMANENTE\\DIRETORIA\\KINEA_INFRA\\07.Pesquisa\\08. Base de debêntures emitidas\\003. Debêntures Incentivadas v01.xlsx", skiprows=2, usecols="B,C,I,J,K,M,N,P,R,T,U,W")
  base_convencional = pandas.read_excel("P:\\PERMANENTE\\DIRETORIA\\KINEA_INFRA\\07.Pesquisa\\08. Base de debêntures emitidas\\004. Debêntures Convencionais v01.xlsx", usecols="B,C,D,E,G,H,I,J,K,L,M,N,O,P,Q")
  ativos_incentivada = Index(base_incentivada["Ativo"])
  ativos_convencional = Index(base_convencional["Ativo"])

def write_in_sheet(sheet, df, header):
  linha = 2
  for row in dataframe_to_rows(df, index = False, header = header):
    coluna = 1
    for celula in row:
      sheet.cell(row = linha, column = coluna).value = celula
      coluna += 1
    linha += 1

def update_base(data):
  global file_path

  downloads_path = os.path.expanduser("~") + "\\Downloads\\"
  # Apagando arquivos antigos
  folder = file_path + "\\base\\"

  for root, dirs, files in os.walk(folder):
    try:
      for f in files:
        os.unlink(os.path.join(root, f))
      for d in dirs:
        shutil.rmtree(os.path.join(root, d))
    except:
      pass
  
  # Salvando novos arquivos

  mes = {1: 'jan', 2: 'fev', 3: 'mar', 4: 'abr', 5: 'mai', 6: 'jun', 7: 'jul', 8: 'ago', 9: 'set', 10: 'out', 11: 'nov', 12: 'dez',}

  nomeima = data.strftime("IMA_%d%m%Y.xls")
  nomeneg = data.strftime("REUNE_Acumulada_%d%m%Y.xls")
  nometax = data.strftime("d%y" + mes[data.month] + "%d.xls")


  shutil.move(downloads_path + nomeima, folder + "imab.xls")
  shutil.move(downloads_path + nomeneg, folder + "reune.xls")
  shutil.move(downloads_path + nometax, folder + "taxas.xls")

def write_tickers(sheet, tickers):
  for linha in range(3, len(tickers)+3):
    sheet.cell(row = linha, column = 3).value = tickers.iloc[linha - 3]["CETIP"]
    sheet.cell(row = linha, column = 8).value = tickers.iloc[linha - 3]["Volume Negociado"]
    sheet.cell(row = linha, column = 9).value = tickers.iloc[linha - 3]["PU Médio"]
    sheet.cell(row = linha, column = 12).value = float(tickers.iloc[linha - 3]["Taxa Média"])/100

def write_in_excel():
  global base_incentivada
  global base_convencional
  global file_path

  imab_complete = pandas.read_html( file_path + "\\base\\imab.xls", header=None , skiprows=1, na_values='-', decimal=",", thousands=".")[0]
  reune_complete = pandas.read_html( file_path + "\\base\\reune.xls", na_values="-", decimal=",", thousands=".")[0]
  taxas_complete = pandas.read_excel( file_path + "\\base\\taxas.xls", sheet_name="IPCA_SPREAD", skiprows=9, skipfooter=4, header=None)
  taxas_complete_di = pandas.read_excel( file_path + "\\base\\taxas.xls", sheet_name="DI_SPREAD", skiprows=9, skipfooter=4, header=None)
  data_ipca = []
  data_di = []
  tickers_ipca = []
  tickers_di = []

  # ---------------------------------------------------- 
  #                  DADOS REUNE  
  for index, row in reune_complete.iterrows():
    if index % 2 == 0 and not (index+1 == len(reune_complete.index)) and not (row[5] == "--"):
      idx = contains_ticker_ipca(row[0]) 
    
      if idx:
        dados = base_incentivada.iloc[idx]

        new_row = [dados["Setor"], dados["Subsetor"], row[0], dados["Titular"], dados["Indexador"],float(row[5]), float(row[6])
                  , float(row[7]), float(row[8]), float(row[9]), float(row[10]), row[11]]

        if dados["Rating (Atual, Br)"] == "" or pandas.isna(dados["Rating (Atual, Br)"]):
          new_row.extend([dados["Rating (Emissão, Br)"], dados["Agência (Emissão, Br)"], dados["Rating equivalente (Emissão, #)"]])
        else:
          new_row.extend([dados["Rating (Atual, Br)"], dados["Agência (Atual, Br)"], dados["Rating equivalente (Atual, #)"]])
        
        data_ipca.append(new_row)
      else:
        idx = contains_ticker_di(row[0])
        
        if idx:
          dados = base_convencional.iloc[idx]

          new_row = [dados["Setor"], dados["Subsetor"], row[0], dados["Titular"], dados["Indexador"], float(row[5]), float(row[6])
                    , float(row[7]), float(row[8]), float(row[9]), float(row[10]), row[11]]

          if dados["Rating (Atual, Br)"] == "" or pandas.isna(dados["Rating (Atual, Br)"]):
            new_row.extend([dados["Rating (Emissão, Br)"], dados["Agência (Emissão, Br)"], dados["Rating equivalente (Emissão, #)"]])
          else:
            new_row.extend([dados["Rating (Atual, Br)"], dados["Agência (Atual, Br)"], dados["Rating equivalente (Atual, #)"]])
          
          data_di.append(new_row)

  data = data_ipca + data_di

  reune_ipca = pandas.DataFrame(data_ipca, columns = ["Setor", "Subsetor", "CETIP", "Titular", "Indexador", "Taxa Mínima", "Taxa Média", "Taxa Máxima", "PU Mínimo", "PU Médio", "PU Máximo", "Volume Negociado", "Rating", "Agência", "Rating Equivalente"])
  reune_ipca = reune_ipca.sort_values(by=["Volume Negociado"], ascending=False) 
  reune_ipca = reune_ipca.sort_values(by=["Setor", "Subsetor", "Rating Equivalente"])
  reune_ipca.reset_index(drop=True, inplace=True)
  tickers_ipca = reune_ipca[["CETIP", "Taxa Média", "PU Médio", "Volume Negociado"]]

  reune_di = pandas.DataFrame(data_di, columns = ["Setor", "Subsetor", "CETIP", "Titular", "Indexador", "Taxa Mínima", "Taxa Média", "Taxa Máxima", "PU Mínimo", "PU Médio", "PU Máximo", "Volume Negociado", "Rating", "Agência", "Rating Equivalente"])
  reune_di = reune_di.sort_values(by=["Volume Negociado"], ascending=False) 
  reune_di = reune_di.sort_values(by=["Setor", "Subsetor", "Rating Equivalente"])
  reune_di.reset_index(drop=True, inplace=True)
  tickers_di = reune_di[["CETIP", "Taxa Média", "PU Médio", "Volume Negociado"]]

  reune = pandas.concat([reune_ipca, reune_di], ignore_index=True)

  # ---------------------------------------------------- 
  #                       IMA-B  
  data = []

  for index, row in imab_complete.iterrows():
    if not (index+1 == len(reune_complete.index)) and not(row[5] == "--") :
      new_row = [row[0], row[1], row[2], row[3], row[4], float(row[5]), float(row[6]), float(row[7]), float(row[8]), float(row[9]), float(row[10]), float(row[11]), float(row[12]), float(row[13]), row[14]
                , row[15], row[16], float(row[17]), float(row[18])]
      data.append(new_row)

  imab = pandas.DataFrame(data, columns = ["Data de Referência", "Títulos", "Data de Vencimento", "Código SELIC", "Código ISIN", "Tx Indicativa", "PU", "PU de Juros", "Quantidade (1.000 títulos)", "Quantidade Teórica"
    , "Carteira a Mercado (R$ mil)", "Peso (%)", "Prazo(d.u.)", "Duration (d.u.)", "Num Op", "Qtd Negociada", "Valor Negociado", "PMR", "Convexidade"])
  
  dia = data[0][0].split("/")
  data_atual = dia[2] + "_" + dia[1] + "_" + dia[0]

  # ---------------------------------------------------- 
  #                  TAXAS ANBIMA  
  data = []

  for index, row in taxas_complete.iterrows():
    idx = contains_ticker_ipca(row[0])
    if idx:
      dados = base_incentivada.iloc[idx]
      new_row = [dados["Subsetor"], "IPCA +"]
      new_row.extend(row[:-2])
      data.append(new_row)

  for index, row in taxas_complete_di.iterrows():
    idx = contains_ticker_di(row[0])
    if idx:
      dados = base_convencional.iloc[idx]
      new_row = [dados["Subsetor"], "CDI +"]
      new_row.extend(row[:-2])
      data.append(new_row)

  taxas = pandas.DataFrame(data, columns = ["Subsetor", "Indexador", "Código", "Nome", "Vencimento", "Índice/Correção", "Taxa de Compra"
  ,"Taxa de Venda", "Taxa Indicativa", "Desvio Padrão", "Taxa Mín", "Taxa Máx", "PU", "% Pu Par", "Duration"])

  wb = load_workbook(filename = file_path + "\\AnaliseSecundarioInfraBase.xlsx")
  ws_reune = wb["Reune"]
  ws_imab = wb["IMA-B"]
  ws_taxas = wb["Taxas Anbima"]
  ws_tabela_ipca = wb["Tabela IPCA"]
  ws_tabela_di = wb["Tabela CDI"]
 
  write_tickers(ws_tabela_ipca, tickers_ipca)
  write_tickers(ws_tabela_di, tickers_di)

  write_in_sheet(ws_reune,reune, False)
  write_in_sheet(ws_imab, imab, False)
  write_in_sheet(ws_taxas,taxas, False)

  print("Salvando arquivo ....")
  wb.save( file_path + "\\AnaliseSecundarioInfra" + data_atual + ".xlsx")
  print("Arquivo salvo")

if __name__ == "__main__":
  print("Iniciando programa..")
  get_path()
  print("atualizar dados? (True/False):")
  wants_update = literal_eval(input())
  
  while wants_update is not True and wants_update is not False:
    wants_update = literal_eval(input("Input não válido! Responda com True ou False:"))

  if wants_update:
    data = input("data (dd/mm/yy):")
    update_base(dt.strptime(data, "%d/%m/%y"))

  read_base_debentures()
  write_in_excel()
  print("end!")

  # teste = send_email("gabriela.barroso-lima@kinea.com.br", "teste", "teste")