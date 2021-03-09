import os
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime
from typing import Tuple

global DATA_DEB
global IMAB
global REUNE
global TAXAS_IPCA
global TAXAS_CDI
global TAXAS_PCT_CDI
global ETTJ
global RATING
global INCENT
global CONVEN
global BDINFRA_path
global date

def clear_sheet(sheet: str) -> None:
    """
        Função para limpar a sheet por completo

        :param sheet: Nome da string para ser limpada no workbook
    """

    global BDINFRA_path
                                                                           
    wb = openpyxl.load_workbook(BDINFRA_path)                                       
    ws = wb[sheet]

    for idx_row, row in enumerate(ws, 1): # Limpando cada célula da sheet
        for col in row:
            col.value = None
    
    wb.save(BDINFRA_path) #Salva as alterações

def pandas_to_excel(sheet: str, df: pd.DataFrame) -> None:
    """
        Realizar a transição de pandas DataFrame para planilha em Excel

        :param sheet: Nome da string para ser populada no workbook
        :param df: DataFrame para ser usada para popular a sheet
    """
    global BDINFRA_path

    wb = openpyxl.load_workbook(BDINFRA_path)                                       
    ws = wb[sheet]

    rows = dataframe_to_rows(df, index=False) #Transforma o dataframe em várias células

    for r_idx, row in enumerate(rows, 1): # Realiza o preenchimento da sheet
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(BDINFRA_path) #Salva as alterações

def date_to_file(dt: str, type: str) -> str:
    """
        Função para converter a data em nome dos files

        :param dt: Data a ser analisada
        :param type: Para qual documento será feito o ajuste
    """

    dt = dt.split('/') # Retira os chars /

    # Criação dos nome dos files de acordo com a data o tipo passado
    if (type == 'IPCA') or (type == 'CDI') or (type == '%CDI'): 
        mes = {1: 'jan', 2: 'fev', 3: 'mar', 4: 'abr', 5: 'mai', 6: 'jun', 7: 'jul', 8: 'ago', 9: 'set', 10: 'out', 11: 'nov', 12: 'dez'}
        file = 'd'+dt[2][-2:]+mes[int(dt[1])]+dt[0]+'.xls'

    if type == 'REUNE':
        file = 'REUNE_Acumulada_'+''.join(dt)+'.csv'

    if type == 'IMAB':
        file = 'IMA_'+''.join(dt)+'.csv'

    if type == 'ETTJ':
        file = 'CurvaZero_'+''.join(dt)+'.csv'
    
    return file # Retorna o nome do file

def clean(df: pd.DataFrame , type: str) -> Tuple[pd.DataFrame, datetime.datetime]:
    """
        Realizar pré-processamentos para bases a serem exploradas

        :param df: DataFrame a ser realizado o pré-processamento
        :param type: Para qual documento será feito o pré-processamento
    """

    # Realiza o pré-processamento a depender do tipo
    if (type == 'IPCA') or (type == 'CDI') or (type == '%CDI'):
        col_1 = list(df.columns)[0]
        idx_srt = list(df.index[df[col_1] == 'Código'])[0]
        idx_end = list(df.index[df[col_1].astype(str).str.contains("[*]")])[0]
        columns = pd.Series(list(df.iloc[idx_srt]))
        col_intind = list(columns[columns.astype(str).str.contains('Intervalo Indicativo')])
        if len(col_intind) == 1:
            columns = columns.fillna('Intervalo Indicativo Máxima')
            columns = columns.replace(col_intind[0], "Intervalo Indicativo Mínimo")
        df.columns = columns
        df = df.iloc[idx_srt+2:idx_end]
        df = df.reset_index(drop=True)
        df['Nome'] = [name[:name.index('*')-2] if '*' in name else name for name in df['Nome']]
        df['Nome'] = [name[:name.index('#')-2] if '#' in name else name for name in df['Nome']]

    if type == 'REUNE':
        df = df.reset_index()
        col_1 = list(df.columns)[0]
        idx_srt = list(df.index[df[col_1].astype(str).str.contains('CETIP')])[0]
        columns = pd.Series(list(df.iloc[idx_srt]))
        df.columns = columns
        df = df.iloc[idx_srt+1:]
        df = df.reset_index(drop=True)
        df = df[df['Agrupamento'] != 'SUBTOTAL']
        df['Taxa Média'] = df['Taxa Média'].str.replace(',', '.')
        df['Preço Médio'] = df['Preço Médio'].str.replace('.', '')
        df['Preço Médio'] = df['Preço Médio'].str.replace(',', '.')

    if type == 'IMAB':
        df = df.reset_index()
        col_1 = list(df.columns)[0]
        idx_srt = list(df.index[df[col_1].astype(str).str.contains('Data')])[0]
        columns = pd.Series(list(df.iloc[idx_srt]))
        columns = [col[:col.index('*')] if '*' in col else col for col in columns]
        df.columns = columns
        df = df.iloc[idx_srt+1:]
        df = df.reset_index(drop=True) 
        df['Duration (d.u.)'] = df['Duration (d.u.)'].str.replace('.', '')
        df['Taxa Indicativa (% a.a.)'] = df['Taxa Indicativa (% a.a.)'].str.replace(',', '.')
    
    if type == 'ETTJ':
        col_1 = list(df.columns)[0]
        col_2 = list(df.columns)[1]
        idx_srt = list(df.index[df[col_1] == 'Vertices'])[0]
        idx_end = list(df.index[df[col_2].isnull()])[2]
        df.columns = list(df.iloc[idx_srt])
        df = df.iloc[idx_srt+1:idx_end]
        df = df.drop(list(set(df.columns) - set(['Vertices', 'ETTJ IPCA', 'ETTJ PREF'])), axis = 1)
        df = df.reset_index(drop=True)
        df['Vertices'] = df['Vertices'].str.replace('.', '')
        df['ETTJ IPCA'] = df['ETTJ IPCA'].str.replace(',', '.')
        df['ETTJ PREF'] = df['ETTJ PREF'].str.replace(',', '.')
    
    if type == 'DATA_DEB':
        df['Codigo do Ativo'] = df['Codigo do Ativo'].str.replace(' ','')
        df.columns = df.columns.str.replace(' ','')

    dt_rat = None
    if type == 'RATING':
        dt_rat = pd.to_datetime(df.columns[0])
        df.columns = df.iloc[0]
        df = df.iloc[1:]

    return df, dt_rat # Retorna o DataFrame já ajustado para análises

def source(dt: str):
    """
        Pegar base de dados crua

        :param str: Data para ser analisada
    """
    
    downloads_path = "~\Downloads"
    downloads_path = os.path.expanduser(downloads_path)

    global date
    date = dt

    global DATA_DEB
    DATA_DEB = pd.read_table(os.path.join(downloads_path, "Debentures.com.br_Caracteristica_em_03-03-2021_as_21-19-42.xls"), encoding='ANSI')
    DATA_DEB, Lixo  = clean(DATA_DEB, 'DATA_DEB')

    global IMAB
    IMAB = pd.read_csv(os.path.join(downloads_path, date_to_file(date, 'IMAB')), sep=";", encoding='ANSI')
    IMAB, Lixo  = clean(IMAB, 'IMAB')

    global REUNE
    REUNE = pd.read_csv(os.path.join(downloads_path, date_to_file(date, 'REUNE')),sep=";",header=2, encoding='ANSI')
    REUNE, Lixo = clean(REUNE, 'REUNE')

    global TAXAS_IPCA
    TAXAS_IPCA = pd.read_excel(os.path.join(downloads_path, date_to_file(date, 'IPCA')), sheet_name='IPCA_SPREAD')
    TAXAS_IPCA, Lixo = clean(TAXAS_IPCA, 'IPCA')

    global TAXAS_CDI
    TAXAS_CDI = pd.read_excel(os.path.join(downloads_path, date_to_file(date, 'CDI')), sheet_name='DI_SPREAD')
    TAXAS_CDI, Lixo = clean(TAXAS_CDI, 'CDI')

    global TAXAS_PCT_CDI
    TAXAS_PCT_CDI = pd.read_excel(os.path.join(downloads_path, date_to_file(date, '%CDI')), sheet_name='DI_PERCENTUAL')
    TAXAS_PCT_CDI, Lixo = clean(TAXAS_PCT_CDI, '%CDI')

    global ETTJ
    ETTJ = pd.read_csv(os.path.join(downloads_path, date_to_file(date, 'ETTJ')),sep=";", encoding='ANSI')
    ETTJ, Lixo = clean(ETTJ, 'ETTJ')

    global RATING
    RATING = pd.read_excel(os.path.join(downloads_path, "Rating.xlsx"))
    RATING, dt_rate = clean(RATING, 'RATING')

    global INCENT
    INCENT = pd.read_excel(os.path.join(downloads_path, "003. Debêntures Incentivadas v01.xlsx"), skiprows=2)

    global CONVEN
    CONVEN = pd.read_excel(os.path.join(downloads_path, "004. Debêntures Convencionais v01.xlsx"))

    global BDINFRA_path 
    BDINFRA_path = os.path.join(downloads_path, "BD Infra - Secundário.xlsx")

    return dt_rate

def fill_sheets(dt_rat) -> None:
    """
        Preenchimento das sheets disponíveis

        :param dt_rat: Data a ser utilizada na sheet de Rating
    """
    print('Início do processo de preenchimento das sheets')

    wis_cadastro()
    wis_rating(dt_rat)
    wis_cadastro_infra()
    wis_ipca_anbima()
    wis_ipca_mercado()
    wis_cdi_anbima()
    wis_cdi_mercado()
    wis_pct_cdi_anbima()
    wis_pct_cdi_mercado()

    print('Processo de preenchimento concluído!')

def wis_cadastro() -> None:
    """
        Função para preenchimento das células na planilha Cadastro
    """

    print('Populando planilha Cadastro...')
    
    global DATA_DEB
    global INCENT
    global CONVEN
    global BDINFRA_path

    sheet = 'Cadastro' # Selecionando a sheet a ser manipulada

    df = pd.read_excel(BDINFRA_path, sheet_name=sheet)
    df = df[df['Ticker'].notna()] # Retira possíveis vazios    

    aux = pd.DataFrame(columns=df.columns)
    for idx_row, row in DATA_DEB.iterrows():

        # Realizando ajustes para determinação se IPCA, CDI ou %CDI
        if (row['indice'] == 'DI') and (row['PercentualMultiplicador/Rentabilidade'] == '100'):
            ind = 'CDI'
        elif (row['indice'] == 'DI') and (row['PercentualMultiplicador/Rentabilidade'] != '100'):
            ind = '%CDI'
        else:
            ind = row['indice']

        # Atualização de linhas do cadastro
        if row['CodigodoAtivo'] in list(df['Ticker']):
            df.loc[df['Ticker'] == row['CodigodoAtivo'], ['Ticker']] = row['CodigodoAtivo']
            df.loc[df['Ticker'] == row['CodigodoAtivo'], ['Emissor']] = row['Empresa']
            df.loc[df['Ticker'] == row['CodigodoAtivo'], ['Índice']] = ind
            df.loc[df['Ticker'] == row['CodigodoAtivo'], ['Data de Saida/Nova Data de Vencimento']] = row['DatadeSaida/NovoVencimento']
            df.loc[df['Ticker'] == row['CodigodoAtivo'], ['Garantia/Especie']] = row['Garantia/Especie']
            df.loc[df['Ticker'] == row['CodigodoAtivo'], ['Valor Nominal na Emissão']] = row['ValorNominalnaEmissao']
            df.loc[df['Ticker'] == row['CodigodoAtivo'], ['Percentual Multiplicador/Rentabilidade']] = row['PercentualMultiplicador/Rentabilidade']
            df.loc[df['Ticker'] == row['CodigodoAtivo'], ['CNPJ']] = row['CNPJ']
            df.loc[df['Ticker'] == row['CodigodoAtivo'], ['Deb. Incentivada (Lei12.431)']] = row['Deb.Incent.(Lei12.431)']
            df.loc[df['Ticker'] == row['CodigodoAtivo'], ['Resgate Antecipado']] = row['ResgateAntecipado']
              
        # Inserção de novos Tickers no cadastro
        else:
            # Determina se o ticker pertecente à infraestrutura ou não
            if (row['CodigodoAtivo'] in list(INCENT['Ativo'])) or (row['CodigodoAtivo'] in list(CONVEN['Ativo'])):
                infra = 1
            else:  
                infra = 0

            # Preenchimento das colunas
            aux = aux.append({
                'Ticker': row['CodigodoAtivo'], 
                'Emissor': row['Empresa'], 
                'Infraestrutura': infra, 
                'Data de Saida/Nova Data de Vencimento': row['DatadeSaida/NovoVencimento'],
                'Garantia/Especie': row['Garantia/Especie'], 
                'Valor Nominal na Emissão': row['ValorNominalnaEmissao'],
                'Índice': ind, 
                'Percentual Multiplicador/Rentabilidade': row['PercentualMultiplicador/Rentabilidade'],
                'CNPJ': row['CNPJ'],
                'Deb. Incentivada (Lei12.431)': row['Deb.Incent.(Lei12.431)'],
                'Resgate Antecipado': row['ResgateAntecipado']}, ignore_index=True)

    df = df.append(aux, ignore_index=True) # Junta os dados novos com os antigos

    # Garatindo que todas as do mesmo emissor estão inclusas em Infraestrutura
    aux = df
    for emissor, df_aux in aux.groupby('Emissor'):
        if (df_aux['Infraestrutura'] == 1).any():
            df.loc[aux['Emissor'] == emissor, ['Infraestrutura']] = 1

    clear_sheet(sheet) # Limpa a sheet com os dados antigos
    pandas_to_excel(sheet, df) # Preenche a sheet com os dados antigos e novos

def wis_cadastro_infra() -> None:
    """
        Função para preenchimento das células na planilha Cadastro Infra
    """

    print('Populando planilha Cadastro Infra...')
    
    global REUNE
    global TAXAS_IPCA
    global TAXAS_CDI
    global INCENT
    global CONVEN
    global BDINFRA_path

    sheet = 'Cadastro Infra' # Selecionando a sheet a ser manipulada

    df_cad = pd.read_excel(BDINFRA_path, sheet_name='Cadastro')
    df_cad = df_cad[df_cad['Infraestrutura'] == 1] # Seleciona as rows somente das que forem de infra

    df_rat = pd.read_excel(BDINFRA_path, sheet_name='Rating')
    df_rat = df_rat.iloc[1:] # Retira as explicações de cada célula

    df = pd.read_excel(BDINFRA_path, sheet_name=sheet)
    df = df[df['Ticker'].notna()] # Retira possíveis vazios

    aux = pd.DataFrame(columns=df.columns)
    for idx_row, row in df_cad.iterrows():
        # Checa se o ticker não está presente na sheet
        if not (row['Ticker'] in list(df['Ticker'])):

            indexador = np.nan
            incentivada = np.nan
            setor = np.nan
            subsetor = np.nan
            ag_rat = np.nan
            rat_emi = np.nan
            # Checa se o ticker é IPCA
            if row['Índice'] == 'IPCA':
                indexador = 'IPCA'
                incentivada = 1
            # Checa se o ticker está é CDI
            elif row['Índice'] == 'CDI': 
                indexador = 'CDI'
                incentivada = 0
            # Checa se o ticker está é PCT_CDI
            elif row['Índice'] == '%CDI': 
                indexador = '%CDI'
                incentivada = 0

            # Checa se o ticker está na planilha INCENTIVADA para preencher o setor e subsetor
            if row['Ticker'] in list(INCENT['Ativo']):
                setor = INCENT[INCENT['Ativo'] == row['Ticker']]['Setor'].iloc[0]
                subsetor = INCENT[INCENT['Ativo'] == row['Ticker']]['Subsetor'].iloc[0]
            # Checa se o ticker está na planilha CONVENCIONAL para preencher o setor e subsetor
            elif row['Ticker'] in list(CONVEN['Ativo']): 
                setor = CONVEN[CONVEN['Ativo'] == row['Ticker']]['Setor'].iloc[0]
                subsetor = CONVEN[CONVEN['Ativo'] == row['Ticker']]['Subsetor'].iloc[0]

            # Checa se o ticker está na planilha Rating para preencher Agência e Rating
            if row['Ticker'] in list(df_rat['Ticker']):
                dt_eval = df_rat[df_rat['Ticker'] == row['Ticker']].sort_values('Dia')['Dia'].iloc[0]
                ag_rat = df_rat[(df_rat['Ticker'] == row['Ticker']) & (df_rat['Dia'] == dt_eval)]['Agência de Rating'].iloc[0]
                rat_emi = df_rat[(df_rat['Ticker'] == row['Ticker']) & (df_rat['Dia'] == dt_eval)]['Rating'].iloc[0]

            # Chega para não colocar uma row vazia
            if (indexador != np.nan):
                # Preenchimento das colunas
                aux = aux.append({
                    'Ticker': row['Ticker'],
                    'Indexador': indexador, 
                    'Incentivada': incentivada,
                    'Emissor': row['Emissor'],
                    'Setor': setor,
                    'Subsetor': subsetor,
                    'Flag Consolidado': np.nan,
                    'Flag Resumo': np.nan,
                    'Agência rating': ag_rat,
                    'Rating de emissão': rat_emi}, ignore_index=True)

    df = df.append(aux, ignore_index=True) # Junta os dados novos com os antigos

    clear_sheet(sheet) # Limpa a sheet com os dados antigos
    pandas_to_excel(sheet, df) # Preenche a sheet com os dados antigos e novos

def wis_rating(dt_rat) -> None:
    """
        Função para preenchimento das células na planilha Rating

        :param dt_rat: Data a ser utilizada na sheet de Rating
    """

    print('Populando planilha Rating...')

    global RATING
    global BDINFRA_path

    sheet = 'Rating' # Selecionando a sheet a ser manipulada

    df_cad = pd.read_excel(BDINFRA_path, sheet_name='Cadastro')
    df_cad = df_cad.iloc[1:] # Retira as explicações de cada célula

    df = pd.read_excel(BDINFRA_path, sheet_name=sheet)
    df = df[df['Ticker'].notna()] # Retira possíveis vazios
    
    aux = pd.DataFrame(columns=df.columns)
    for idx_row, row in df_cad.iterrows():
        # Checa se o ticker não está presente na sheet ou se, estando, está sendo uma data diferente para popular
        if (not (row['Ticker'] in list(df['Ticker']))) or (not (dt_rat in list(df[df['Ticker'] == row['Ticker']]['Dia']))):
            ag_ret = np.nan
            rat = np.nan
            rat_eq = np.nan
            # Checa se o ticker está nas planilha RATING
            if row['Ticker'] in list(RATING['Emissor']):
                ag_ret = RATING[RATING['Emissor'] == row['Ticker']]['Agência'].iloc[0]
                rat = RATING[RATING['Emissor'] == row['Ticker']]['Rating'].iloc[0]
                rat_eq = RATING[RATING['Emissor'] == row['Ticker']]['Rating Equivalente'].iloc[0]
                # Preenchimento das colunas
                aux = aux.append({
                    'Dia': dt_rat,
                    'Ticker': row['Ticker'], 
                    'Agência de Rating': ag_ret,
                    'Rating': rat,
                    'Rating Equivalente': rat_eq}, ignore_index=True)

				
    df = df.append(aux, ignore_index=True) # Junta os dados novos com os antigos

    clear_sheet(sheet) # Limpa a sheet com os dados antigos
    pandas_to_excel(sheet, df) # Preenche a sheet com os dados antigos e novos

def wis_ipca_anbima() -> None:
    """
        Função para preenchimento das células na planilha IPCA - ANBIMA
    """

    print('Populando planilha IPCA - ANBIMA...')

    global TAXAS_IPCA
    global BDINFRA_path
    global IMAB
    global ETTJ
    global date

    sheet = 'IPCA - ANBIMA' # Selecionando a sheet a ser manipulada

    df_cad = pd.read_excel(BDINFRA_path, sheet_name='Cadastro')
    df_cad = df_cad[df_cad['Índice'] == 'IPCA'] # Seleciona as rows somente das que forem com o indexador em IPCA

    df = pd.read_excel(BDINFRA_path, sheet_name=sheet)
    df = df[df['Ticker'].notna()] # Retira possíveis vazios

    aux = pd.DataFrame(columns=df.columns)
    for idx_row, row in df_cad.iterrows(): 
        # Checa se o ticker não está presente na sheet ou se, estando, está sendo uma data diferente para popular
        if (not (row['Ticker'] in list(df['Ticker']))) or (not (date in list(df[df['Ticker'] == row['Ticker']]['Dia']))):
            
            ticker = np.nan
            pu_anbima = np.nan
            durat_anbima = np.nan
            taxa_anbima = np.nan
            ntnb_ref_anbima_date = np.nan
            ntnb_ref_anbima_tax = np.nan
            spread_anbima_ntnb = np.nan
            ettj_ref_anbima_tax = np.nan
            spread_anbima_ettj = np.nan 
            # Checa se o ticker está nas planilha TAXAS_IPCA
            if row['Ticker'] in list(TAXAS_IPCA['Código']):
                ticker = row['Ticker']
                pu_anbima = TAXAS_IPCA[TAXAS_IPCA['Código'] == row['Ticker']]['PU'].iloc[0]
                pu_anbima = pu_anbima if pu_anbima != 'N/D' else np.nan
                taxa_anbima = TAXAS_IPCA[TAXAS_IPCA['Código'] == row['Ticker']]['Taxa Indicativa'].iloc[0]
                durat_anbima = TAXAS_IPCA[TAXAS_IPCA['Código'] == row['Ticker']]['Duration'].iloc[0] 
                durat_anbima = int(durat_anbima) if durat_anbima != 'N/D' else np.nan
                # Caso tenha um valor para duration
                if durat_anbima != np.nan:  
                    # Cálculo do valor mais prócimo de duration disponível
                    durat_ref_ntnb = np.array(IMAB['Duration (d.u.)'])[(np.abs(np.array(IMAB['Duration (d.u.)']).astype(int) - durat_anbima)).argmin()] 
                    # Filta a data de vencimento correspondente
                    ntnb_ref_anbima_date = IMAB[IMAB['Duration (d.u.)'] == durat_ref_ntnb]['Data  de Vencimento'].iloc[0]
                    # Filta a taxa indicativa correspondente
                    ntnb_ref_anbima_tax = float(IMAB[IMAB['Duration (d.u.)'] == durat_ref_ntnb]['Taxa Indicativa (% a.a.)'].iloc[0])
                    # Cálculo do spread entre anbima e ntnb
                    spread_anbima_ntnb = (1+taxa_anbima)/(1+ntnb_ref_anbima_tax) - 1
                    # Interpolação entre os durations de ettj e suas taxas ettj ipca e o duration usado
                    ettj_ref_anbima_tax = np.interp(durat_anbima, np.array(ETTJ['Vertices']).astype(int), np.array(ETTJ['ETTJ IPCA']).astype(float))
                    # Cálculo do spread entre anbima e ettj
                    spread_anbima_ettj = (1+taxa_anbima)/(1+ettj_ref_anbima_tax) - 1
                # Preenchimento das colunas
                aux = aux.append({
                    'Ticker': ticker,
                    'Dia': date, 
                    'PU ANBIMA': pu_anbima,
                    'Duration ANBIMA': durat_anbima,
                    'Taxa ANBIMA': taxa_anbima,
                    'NTN-B Ref ANBIMA': ntnb_ref_anbima_date,
                    'Taxa NTN-B Ref ANBIMA': ntnb_ref_anbima_tax,
                    'Spread B Ref ANBIMA': spread_anbima_ntnb,
                    'Taxa ETTJ Ref ANBIMA': ettj_ref_anbima_tax,
                    'Spread ETTJ Ref ANBIMA': spread_anbima_ettj}, ignore_index=True)

    df = df.append(aux, ignore_index=True) # Junta os dados novos com os antigos

    clear_sheet(sheet) # Limpa a sheet com os dados antigos
    pandas_to_excel(sheet, df) # Preenche a sheet com os dados antigos e novos

def wis_ipca_mercado() -> None:
    """
        Função para preenchimento das células na planilha IPCA - Mercado
    """

    print('Populando planilha IPCA - Mercado...')

    global REUNE
    global IMAB
    global ETTJ
    global BDINFRA_path
    global date

    sheet = 'IPCA - Mercado' # Selecionando a sheet a ser manipulada
    
    df_cad = pd.read_excel(BDINFRA_path, sheet_name='Cadastro')
    df_cad = df_cad[df_cad['Índice'] == 'IPCA'] # Seleciona as rows somente das que forem com o indexador em IPCA

    df = pd.read_excel(BDINFRA_path, sheet_name=sheet)
    df = df[df['Ticker'].notna()] # Retira possíveis vazios

    aux = pd.DataFrame(columns=df.columns)
    for idx_row, row in df_cad.iterrows():
        # Checa se o ticker não está presente na sheet ou se, estando, está sendo uma data diferente para popular 
        if (not (row['Ticker'] in list(df['Ticker']))) or (not (date in list(df[df['Ticker'] == row['Ticker']]['Dia']))):
            
            ticker = np.nan
            vol_neg = np.nan
            pu_mercado = np.nan
            durat_mercado = np.nan
            taxa_mercado = np.nan
            ntnb_ref_mercado_date = np.nan
            ntnb_ref_mercado_tax = np.nan
            spread_mercado_ntnb = np.nan
            ettj_ref_mercado_tax = np.nan
            spread_mercado_ettj = np.nan 
            # Checa se o ticker está nas planilha REUNE
            if row['Ticker'] in list(REUNE['CETIP']):
                ticker = row['Ticker']
                vol_neg =  REUNE[REUNE['CETIP'] == row['Ticker']]['Faixa de Volume'].iloc[0]
                pu_mercado = float(REUNE[REUNE['CETIP'] == row['Ticker']]['Preço Médio'].iloc[0])
                taxa_mercado = REUNE[REUNE['CETIP'] == row['Ticker']]['Taxa Média'].iloc[0]
                taxa_mercado = float(taxa_mercado) if taxa_mercado != '--' else np.nan
                # Por enquanto, utilizando o duration da ANBIMA
                durat_mercado = TAXAS_IPCA[TAXAS_IPCA['Código'] == row['Ticker']]['Duration'].iloc[0] if row['Ticker'] in list(TAXAS_IPCA['Código']) else 'N/D'
                durat_mercado = int(durat_mercado) if durat_mercado != 'N/D' else np.nan 
                # Caso tenha um valor para duration e de taxa indicativa
                if (durat_mercado != np.nan) and (taxa_mercado != np.nan):
                    # Cálculo do valor mais prócimo de duration disponível
                    durat_ref_ntnb = np.array(IMAB['Duration (d.u.)'])[(np.abs(np.array(IMAB['Duration (d.u.)']).astype(int) - durat_mercado)).argmin()]
                    # Filta a data de vencimento correspondente
                    ntnb_ref_mercado_date = IMAB[IMAB['Duration (d.u.)'] == durat_ref_ntnb]['Data  de Vencimento'].iloc[0]
                    # Filta a taxa indicativa correspondente
                    ntnb_ref_mercado_tax = float(IMAB[IMAB['Duration (d.u.)'] == durat_ref_ntnb]['Taxa Indicativa (% a.a.)'].iloc[0])
                    # Cálculo do spread entre mercado e ntnb
                    spread_mercado_ntnb = (1+taxa_mercado)/(1+ntnb_ref_mercado_tax) - 1
                    # Interpolação entre os durations de ettj e suas taxas ettj ipca e o duration usado
                    ettj_ref_mercado_tax = np.interp(durat_mercado, np.array(ETTJ['Vertices']).astype(int), np.array(ETTJ['ETTJ IPCA']).astype(float))
                    # Cálculo do spread entre mercado e ettj
                    spread_mercado_ettj = (1+taxa_mercado)/(1+ettj_ref_mercado_tax) - 1
                # Preenchimento das colunas
                aux = aux.append({
                    'Ticker': ticker,
                    'Dia': date, 
                    'Volume Negociado': vol_neg,
                    'PU Mercado': pu_mercado,
                    'Duration Mercado': durat_mercado,
                    'Taxa Mercado': taxa_mercado,
                    'NTN-B Ref Mercado': ntnb_ref_mercado_date,
                    'Taxa NTN-B Ref Mercado': ntnb_ref_mercado_tax,
                    'Spread B Ref Mercado': spread_mercado_ntnb,
                    'Taxa ETTJ Ref Mercado': ettj_ref_mercado_tax,
                    'Spread ETTJ Ref Mercado ': spread_mercado_ettj}, ignore_index=True)

    df = df.append(aux, ignore_index=True) # Junta os dados novos com os antigos

    clear_sheet(sheet) # Limpa a sheet com os dados antigos
    pandas_to_excel(sheet, df) # Preenche a sheet com os dados antigos e novos

def wis_cdi_anbima() -> None:
    """
        Função para preenchimento das células na planilha CDI - ANBIMA
    """

    print('Populando planilha CDI - ANBIMA...')

    global TAXAS_CDI
    global BDINFRA_path
    global date

    sheet = 'CDI - ANBIMA' # Selecionando a sheet a ser manipulada

    df_cad = pd.read_excel(BDINFRA_path, sheet_name='Cadastro')
    df_cad = df_cad[df_cad['Índice'] == 'CDI'] # Seleciona as rows somente das que forem com o indexador em CDI

    df = pd.read_excel(BDINFRA_path, sheet_name=sheet)
    df = df[df['Ticker'].notna()] # Retira possíveis vazios

    aux = pd.DataFrame(columns=df.columns)
    for idx_row, row in df_cad.iterrows(): 
        # Checa se o ticker não está presente na sheet ou se, estando, está sendo uma data diferente para popular
        if (not (row['Ticker'] in list(df['Ticker']))) or (not (date in list(df[df['Ticker'] == row['Ticker']]['Dia']))):

            ticker = np.nan
            pu_anbima = np.nan
            durat_anbima = np.nan
            taxa_anbima = np.nan
            spread_anbima = np.nan
            # Checa se o ticker está nas planilha TAXAS_CDI
            if row['Ticker'] in list(TAXAS_CDI['Código']):
                ticker = row['Ticker']
                pu_anbima = TAXAS_CDI[TAXAS_CDI['Código'] == row['Ticker']]['PU'].iloc[0]
                pu_anbima = pu_anbima if pu_anbima != 'N/D' else np.nan
                durat_anbima = TAXAS_CDI[TAXAS_CDI['Código'] == row['Ticker']]['Duration'].iloc[0]
                durat_anbima = int(durat_anbima) if durat_anbima != 'N/D' else np.nan
                taxa_anbima = TAXAS_CDI[TAXAS_CDI['Código'] == row['Ticker']]['Taxa Indicativa'].iloc[0]
                spread_anbima = taxa_anbima
                # Preenchimento das colunas
                aux = aux.append({
                    'Ticker': ticker,
                    'Dia': date, 
                    'PU ANBIMA': pu_anbima,
                    'Duration ANBIMA': durat_anbima,
                    'Taxa ANBIMA': taxa_anbima,
                    'Spread ANBIMA': spread_anbima}, ignore_index=True)

    df = df.append(aux, ignore_index=True) # Junta os dados novos com os antigos

    clear_sheet(sheet) # Limpa a sheet com os dados antigos
    pandas_to_excel(sheet, df) # Preenche a sheet com os dados antigos e novos

def wis_cdi_mercado() -> None:
    """
        Função para preenchimento das células na planilha CDI - Mercado
    """

    print('Populando planilha CDI - Mercado...')

    global REUNE
    global IMAB
    global BDINFRA_path
    global date

    sheet = 'CDI - Mercado' # Selecionando a sheet a ser manipulada

    df_cad = pd.read_excel(BDINFRA_path, sheet_name='Cadastro')
    df_cad = df_cad[df_cad['Índice'] == 'CDI'] # Seleciona as rows somente das que forem com o indexador em CDI

    df = pd.read_excel(BDINFRA_path, sheet_name=sheet)
    df = df[df['Ticker'].notna()] # Retira possíveis vazios

    aux = pd.DataFrame(columns=df.columns)
    for idx_row, row in df_cad.iterrows(): 
        # Checa se o ticker não está presente na sheet ou se, estando, está sendo uma data diferente para popular
        if (not (row['Ticker'] in list(df['Ticker']))) or (not (date in list(df[df['Ticker'] == row['Ticker']]['Dia']))):

            ticker = np.nan
            vol_neg = np.nan
            pu_mercado = np.nan
            durat_mercado = np.nan
            taxa_mercado = np.nan
            spread_mercado = np.nan
            # Checa se o ticker está nas planilha REUNE
            if row['Ticker'] in list(REUNE['CETIP']):
                ticker = row['Ticker']
                vol_neg = REUNE[REUNE['CETIP'] == row['Ticker']]['Faixa de Volume'].iloc[0]
                pu_mercado = float(REUNE[REUNE['CETIP'] == row['Ticker']]['Preço Médio'].iloc[0])
                taxa_mercado = REUNE[REUNE['CETIP'] == row['Ticker']]['Taxa Média'].iloc[0]
                taxa_mercado = float(taxa_mercado) if taxa_mercado != '--' else np.nan
                spread_mercado = taxa_mercado
                # Caso tenha um valor de taxa indicativa
                if (taxa_mercado != np.nan):
                    # Interpolação entre as taxas de ntnb e os durations e a taxa usada
                    durat_mercado = np.interp(taxa_mercado, np.array(IMAB['Taxa Indicativa (% a.a.)']).astype(float), np.array(IMAB['Duration (d.u.)']).astype(int))
                # Preenchimento das colunas
                aux = aux.append({
                    'Ticker': ticker,
                    'Dia': date, 
                    'Volume Negociado': vol_neg,
                    'PU Mercado': pu_mercado,
                    'Duration Mercado': durat_mercado,
                    'Taxa Mercado': taxa_mercado,
                    'Spread Mercado': spread_mercado}, ignore_index=True)

    df = df.append(aux, ignore_index=True) # Junta os dados novos com os antigos

    clear_sheet(sheet) # Limpa a sheet com os dados antigos
    pandas_to_excel(sheet, df) # Preenche a sheet com os dados antigos e novos

def wis_pct_cdi_anbima() -> None: 
    """
        Função para preenchimento das células na planilha %CDI - ANBIMA
    """

    print('Populando planilha %CDI - ANBIMA...')

    global TAXAS_PCT_CDI
    global BDINFRA_path
    global date

    sheet = '%CDI - ANBIMA' # Selecionando a sheet a ser manipulada

    df_cad = pd.read_excel(BDINFRA_path, sheet_name='Cadastro')
    df_cad = df_cad[df_cad['Índice'] == '%CDI'] # Seleciona as rows somente das que forem com o indexador em %CDI

    df = pd.read_excel(BDINFRA_path, sheet_name=sheet)
    df = df[df['Ticker'].notna()] # Retira possíveis vazios

    aux = pd.DataFrame(columns=df.columns)
    for idx_row, row in df_cad.iterrows(): 
        # Checa se o ticker não está presente na sheet ou se, estando, está sendo uma data diferente para popular
        if (not (row['Ticker'] in list(df['Ticker']))) or (not (date in list(df[df['Ticker'] == row['Ticker']]['Dia']))):

            ticker = np.nan
            taxa_emi = np.nan
            pu_anbima = np.nan
            durat_anbima = np.nan
            taxa_anbima = np.nan
            pre_ref_anbima_date = np.nan
            pre_ref_anbima_tax = np.nan
            spread_anbima_pre = np.nan
            ettj_ref_anbima_tax = np.nan
            spread_anbima_ettj = np.nan

            # Checa se o ticker está nas planilha TAXAS_PCT_CDI
            if row['Ticker'] in list(TAXAS_PCT_CDI['Código']):
                ticker = row['Ticker']
                taxa_emi = row['Percentual Multiplicador/Rentabilidade']
                pu_anbima = TAXAS_PCT_CDI[TAXAS_PCT_CDI['Código'] == row['Ticker']]['PU'].iloc[0]
                pu_anbima = pu_anbima if pu_anbima != 'N/D' else np.nan 
                durat_anbima = TAXAS_PCT_CDI[TAXAS_PCT_CDI['Código'] == row['Ticker']]['Duration'].iloc[0]
                durat_anbima = int(durat_anbima) if durat_anbima != 'N/D' else np.nan
                taxa_anbima = TAXAS_PCT_CDI[TAXAS_PCT_CDI['Código'] == row['Ticker']]['Taxa Indicativa'].iloc[0]
                # Caso tenha um valor de duration
                if durat_anbima != np.nan:
                    # Valores nulos por enquanto
                    pre_ref_anbima_date = np.nan
                    pre_ref_anbima_tax = np.nan
                    spread_anbima_pre = np.nan
                    # Interpolação entre os durations de ettj e seus taxas ettj pref e o duration usado
                    ettj_ref_anbima_tax = np.interp(durat_anbima, np.array(ETTJ['Vertices']).astype(int), np.array(ETTJ['ETTJ PREF']).astype(float))
                    # Cálculo do spread entre anbima e ettj
                    spread_anbima_ettj = (1+taxa_anbima)/(1+ettj_ref_anbima_tax) - 1
                # Preenchimento das colunas
                aux = aux.append({
                    'Ticker': ticker,
                    'Dia': date, 
                    'Taxa Emissão': taxa_emi,
                    'PU ANBIMA': pu_anbima,
                    'Duration ANBIMA': durat_anbima,
                    'Taxa ANBIMA': taxa_anbima,
                    'Pré Ref ANBIMA': pre_ref_anbima_date,
                    'Taxa Pré Ref ANBIMA': pre_ref_anbima_tax,
                    'Spread Pré Ref ANBIMA': spread_anbima_pre,
                    'Taxa ETTJ Ref ANBIMA': ettj_ref_anbima_tax,
                    'Spread ETTJ Ref ANBIMA': spread_anbima_ettj}, ignore_index=True)

    df = df.append(aux, ignore_index=True) # Junta os dados novos com os antigos

    clear_sheet(sheet) # Limpa a sheet com os dados antigos
    pandas_to_excel(sheet, df) # Preenche a sheet com os dados antigos e novos

def wis_pct_cdi_mercado() -> None:
    """
        Função para preenchimento das células na planilha %CDI - Mercado
    """

    print('Populando planilha %CDI - Mercado...')

    global REUNE
    global BDINFRA_path
    global date

    sheet = '%CDI - Mercado' # Selecionando a sheet a ser manipulada

    df_cad = pd.read_excel(BDINFRA_path, sheet_name='Cadastro')
    df_cad = df_cad[df_cad['Índice'] == '%CDI'] # Seleciona as rows somente das que forem com o indexador em %CDI

    df = pd.read_excel(BDINFRA_path, sheet_name=sheet)
    df = df[df['Ticker'].notna()] # Retira possíveis vazios

    aux = pd.DataFrame(columns=df.columns)
    for idx_row, row in df_cad.iterrows(): 
        # Checa se o ticker não está presente na sheet ou se, estando, está sendo uma data diferente para popular
        if (not (row['Ticker'] in list(df['Ticker']))) or (not (date in list(df[df['Ticker'] == row['Ticker']]['Dia']))):

            ticker = np.nan
            taxa_emi = np.nan
            vol_neg = np.nan
            pu_mercado = np.nan
            durat_mercado = np.nan
            taxa_mercado = np.nan
            pre_ref_mercado_date = np.nan
            pre_ref_mercado_tax = np.nan
            spread_mercado_pre = np.nan
            ettj_ref_mercado_tax = np.nan
            spread_mercado_ettj = np.nan
            # Checa se o ticker está nas planilha REUNE
            if row['Ticker'] in list(REUNE['CETIP']):
                ticker = row['Ticker']
                taxa_emi = row['Percentual Multiplicador/Rentabilidade']
                vol_neg = REUNE[REUNE['CETIP'] == row['Ticker']]['Faixa de Volume'].iloc[0]
                pu_mercado = float(REUNE[REUNE['CETIP'] == row['Ticker']]['Preço Médio'].iloc[0])
                durat_mercado = TAXAS_PCT_CDI[TAXAS_PCT_CDI['Código'] == row['Ticker']]['Duration'].iloc[0] if row['Ticker'] in list(TAXAS_PCT_CDI['Código']) else 'N/D'
                durat_mercado = int(durat_mercado) if durat_mercado != 'N/D' else np.nan 
                taxa_mercado = REUNE[REUNE['CETIP'] == row['Ticker']]['Taxa Média'].iloc[0]
                taxa_mercado = float(taxa_mercado) if taxa_mercado != '--' else np.nan
                # Caso tenha um valor para duration e de taxa indicativa
                if (durat_mercado != np.nan) and (taxa_mercado != np.nan):
                    # Valores nulos por enquanto
                    pre_ref_mercado_date = np.nan
                    pre_ref_mercado_tax = np.nan
                    spread_mercado_pre = np.nan
                    # Interpolação entre os durations de ettj e seus taxas ettj pref e o duration usado
                    ettj_ref_mercado_tax = np.interp(durat_mercado, np.array(ETTJ['Vertices']).astype(int), np.array(ETTJ['ETTJ PREF']).astype(float))
                    # Cálculo do spread entre mercado e ettj
                    spread_mercado_ettj = (1+taxa_mercado)/(1+ettj_ref_mercado_tax) - 1
                # Preenchimento das colunas
                aux = aux.append({
                    'Ticker': ticker,
                    'Dia': date, 
                    'Taxa Emissão': taxa_emi,
                    'Volume Negociado': vol_neg, 
                    'PU Mercado': pu_mercado,
                    'Duration Mercado': durat_mercado,
                    'Taxa Mercado': taxa_mercado,
                    'Pré Ref Mercado': pre_ref_mercado_date,
                    'Taxa Pré Ref Mercado': pre_ref_mercado_tax,
                    'Spread Pré Ref Mercado': spread_mercado_pre,
                    'Taxa ETTJ Ref Mercado': ettj_ref_mercado_tax,
                    'Spread ETTJ Ref Mercado': spread_mercado_ettj}, ignore_index=True)

    df = df.append(aux, ignore_index=True) # Junta os dados novos com os antigos

    clear_sheet(sheet) # Limpa a sheet com os dados antigos
    pandas_to_excel(sheet, df) # Preenche a sheet com os dados antigos e novos