import os
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime
from typing import Tuple

global DATA_DEB
global RATING
global CLASS_DEB
global BDINFRA_path
global downloads_path
global dt_rat

def clear_sheet(sheet: str) -> None:
    """
        Função para limpar a sheet por completo

        :param sheet: Nome da string para ser limpada no workbook
    """

    global BDINFRA_path
                                                                           
    wb = openpyxl.load_workbook(BDINFRA_path)                                       
    ws = wb[sheet]

    for idx_row, row in enumerate(ws, 1): # Limpando cada célula da sheet
        for col in row:
            col.value = None
    
    wb.save(BDINFRA_path) #Salva as alterações

def pandas_to_excel(sheet: str, df: pd.DataFrame) -> None:
    """
        Realizar a transição de pandas DataFrame para planilha em Excel

        :param sheet: Nome da string para ser populada no workbook
        :param df: DataFrame para ser usada para popular a sheet
    """
    global BDINFRA_path

    wb = openpyxl.load_workbook(BDINFRA_path)                                       
    ws = wb[sheet]

    rows = dataframe_to_rows(df, index=False) #Transforma o dataframe em várias células

    for r_idx, row in enumerate(rows, 1): # Realiza o preenchimento da sheet
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(BDINFRA_path) #Salva as alterações

def date_to_file(dt: str, type: str, sem: str) -> str:
    """
        Função para converter a data em nome dos files

        :param dt: Data a ser analisada
        :param type: Para qual documento será feito o ajuste
        :param sem: Semestre a ser analisado
    """

    dt = dt.split('/') # Retira os chars /

    # Criação dos nome dos files de acordo com a data o tipo passado
    if type == 'TAXAS': 
        file = 'taxas - '+ dt[2] + '_' + sem + '.xlsx'

    if type == 'REUNE':
        file = 'REUNE_Acumulada_'+''.join(dt)+'.csv'

    if type == 'IMAB':
        file = 'imab - '+ dt[2] + '_' + sem + '.xlsx'

    if type == 'ETTJ':
        file = 'ettj - '+ dt[2] + '_' + sem + '.xlsx'

    if type == 'Pre_REF':
        file = 'Pre Ref - '+ dt[2] + '_' + sem + '.xlsx'
    
    return file # Retorna o nome do file

def clean(df: pd.DataFrame , type: str, date: str) -> Tuple[pd.DataFrame, datetime.datetime]:
    """
        Realizar pré-processamentos para bases a serem exploradas

        :param df: DataFrame a ser realizado o pré-processamento
        :param type: Para qual documento será feito o pré-processamento
        :param date: Data para ser analisada
    """

    # Realiza o pré-processamento a depender do tipo

    if type == 'TAXAS':
        df.rename(columns={ df.columns[0]: "Tickers", df.columns[1]: "Tipo"}, inplace = True)
        df.loc[:, ['Tickers']] = df.loc[:, ['Tickers']].ffill(axis=0)
        df = df.set_index(['Tickers', 'Tipo'])
        dt = date.split('/')
        df = df[pd.to_datetime(dt[2] + '-' + dt[1] + '-' + dt[0])]
        tikcers_notnan = df.notna().groupby(level=0).any().where(lambda x : x == True).dropna().keys()
        df = df[df.index.get_level_values(level=0).isin(tikcers_notnan)]

    if type == 'IMAB':
        df.rename(columns={ df.columns[0]: "Tickers", df.columns[1]: "Tipo"}, inplace = True)
        df.loc[:, ['Tickers']] = df.loc[:, ['Tickers']].ffill(axis=0)
        df = df.set_index(['Tickers', 'Tipo'])
        DataVenc = [idx[-10:][-2:]+ '/' + idx[-10:][5:7] + '/' + idx[-10:][:4] for idx in df.index.get_level_values(level=0)]
        df = df.reset_index()
        df['Tickers'] = [tck[:-13] for tck in df['Tickers']]
        df['Data de Vencimento'] = DataVenc
        df = df.set_index(['Tickers','Data de Vencimento', 'Tipo'])
        dt = date.split('/')
        df = df[pd.to_datetime(dt[2] + '-' + dt[1] + '-' + dt[0])]
        tikcers_notnan = df.notna().groupby(level=1).any().where(lambda x : x == True).dropna().keys()
        df = df[df.index.get_level_values(level=1).isin(tikcers_notnan)]

    if type == 'REUNE':
        df = df.reset_index()
        col_1 = list(df.columns)[0]
        idx_srt = list(df.index[df[col_1].astype(str).str.contains('CETIP')])[0]
        columns = pd.Series(list(df.iloc[idx_srt]))
        df.columns = columns
        df = df.iloc[idx_srt+1:]
        df = df.reset_index(drop=True)
        df = df[df['Agrupamento'] != 'SUBTOTAL']
        df['Taxa Média'] = df['Taxa Média'].str.replace(',', '.', regex=True)
        df['Preço Médio'] = df['Preço Médio'].str.replace('.', '', regex=True)
        df['Preço Médio'] = df['Preço Médio'].str.replace(',', '.', regex=True)

    if type == 'ETTJ':
        df.rename(columns={ df.columns[0]: "Tickers", df.columns[1]: "Tipo"}, inplace = True)
        df.loc[:, ['Tickers']] = df.loc[:, ['Tickers']].ffill(axis=0)
        df = df.set_index(['Tickers', 'Tipo'])
        df = df[df.index.get_level_values(level=0).str.contains('ETTJ')]
        durs = [int(idx[-4:]) for idx in df.index.get_level_values(level=0)]
        df = df.reset_index()
        df['Tickers'] = [tck[:-5] for tck in df['Tickers']]
        df = df.drop('Tipo', axis=1)
        df['Duration'] = durs
        df = df.set_index(['Tickers','Duration'])
        dt = date.split('/')
        df = df[pd.to_datetime(dt[2] + '-' + dt[1] + '-' + dt[0])]
        dur_notnan = df.notna().groupby(level=['Tickers', 'Duration']).any().where(lambda x : x == True).dropna().keys()
        df = df[df.index.isin(dur_notnan)]

    if type == 'DATA_DEB':
        df['Codigo do Ativo'] = df['Codigo do Ativo'].str.replace(' ','')
        df.columns = df.columns.str.replace(' ','')

    if type == 'Feriados':
        df = list(df[0])
        df = [str(d.date()) for d in df]
    
    if type == 'Pre_REF':
        df = df.rename(columns = {df.columns[0] : 'Data'})
        df = df.iloc[1:]

    dt = None
    if type == 'RATING':
        dt = pd.to_datetime(df.columns[0])
        df.columns = df.iloc[0]
        df = df.iloc[1:]

    return df, dt # Retorna o DataFrame já ajustado para análises

def source():
    """
        Pegar base de dados crua

    """

    global downloads_path
    downloads_path = "~\Downloads"
    downloads_path = os.path.expanduser(downloads_path)

    global DATA_DEB
    DATA_DEB = pd.read_table(os.path.join(downloads_path, "Debentures.com.br.xls"), encoding='ANSI', skiprows=3)
    DATA_DEB, Lixo  = clean(DATA_DEB, 'DATA_DEB', None)

    global RATING
    global dt_rat
    RATING = pd.read_excel(os.path.join(downloads_path, "Rating.xlsx"))
    RATING, dt_rat = clean(RATING, 'RATING', None)

    global Feriados
    Feriados = pd.read_excel(os.path.join(downloads_path, "Feriados.xlsx"), header=None)
    Feriados, Lixo = clean(Feriados, 'Feriados', None)

    global CLASS_DEB
    CLASS_DEB = pd.read_excel(os.path.join(downloads_path, "Classificação Debêntures.xlsx"))

    global BDINFRA_path 
    BDINFRA_path = os.path.join(downloads_path, "BD Infra - Secundário - Histórico.xlsx")

def wis_cadastro() -> None:
    """
        Função para preenchimento das células na planilha Cadastro
    """

    print('Populando planilha Cadastro...')
    
    global DATA_DEB
    global CLASS_DEB
    global BDINFRA_path

    sheet = 'Cadastro' # Selecionando a sheet a ser manipulada

    df = pd.read_excel(BDINFRA_path, sheet_name=sheet)
    df = df[df['Ticker'].notna()] # Retira possíveis vazios    

    aux = pd.DataFrame(columns=df.columns)
    for idx_row, row in DATA_DEB.iterrows():

        # Realizando ajustes para determinação se IPCA, CDI ou %CDI
        if (row['indice'] == 'DI') and (row['PercentualMultiplicador/Rentabilidade'] == '100'):
            ind = 'CDI'
        elif (row['indice'] == 'DI') and (row['PercentualMultiplicador/Rentabilidade'] != '100'):
            ind = '%CDI'
        else:
            ind = row['indice']

        perct = row['PercentualMultiplicador/Rentabilidade'] if row['PercentualMultiplicador/Rentabilidade'] != ' -' else np.nan

        # Atualização de linhas do cadastro
        if row['CodigodoAtivo'] in list(df['Ticker']):
            df.loc[df['Ticker'] == row['CodigodoAtivo'], ['Ticker']] = row['CodigodoAtivo']
            df.loc[df['Ticker'] == row['CodigodoAtivo'], ['Emissor']] = row['Empresa']
            df.loc[df['Ticker'] == row['CodigodoAtivo'], ['Índice']] = ind
            df.loc[df['Ticker'] == row['CodigodoAtivo'], ['Data de Saida/Nova Data de Vencimento']] = row['DatadeSaida/NovoVencimento']
            df.loc[df['Ticker'] == row['CodigodoAtivo'], ['Garantia/Especie']] = row['Garantia/Especie']
            df.loc[df['Ticker'] == row['CodigodoAtivo'], ['Valor Nominal na Emissão']] = row['ValorNominalnaEmissao']
            df.loc[df['Ticker'] == row['CodigodoAtivo'], ['Quantidade no Mercado']] = row['QuantidadeemMercado']
            df.loc[df['Ticker'] == row['CodigodoAtivo'], ['Percentual Multiplicador/Rentabilidade']] = perct
            df.loc[df['Ticker'] == row['CodigodoAtivo'], ['CNPJ']] = row['CNPJ']
            df.loc[df['Ticker'] == row['CodigodoAtivo'], ['Deb. Incentivada (Lei12.431)']] = row['Deb.Incent.(Lei12.431)']
            df.loc[df['Ticker'] == row['CodigodoAtivo'], ['Resgate Antecipado']] = row['ResgateAntecipado']
              
        # Inserção de novos Tickers no cadastro
        else:
            # Determina se o ticker pertecente à infraestrutura ou não e à kinea ou não
            if row['Empresa'] in list(CLASS_DEB['Emissor']):
                infra = CLASS_DEB[CLASS_DEB['Emissor'] == row['Empresa']]['Infraestrutura'].iloc[0]
            else:  
                infra = 0

            perct = row['PercentualMultiplicador/Rentabilidade'] if row['PercentualMultiplicador/Rentabilidade'] != ' -' else np.nan
            kinea = 0 # Por default, assume que não está presente na Kinea    

            # Preenchimento das colunas
            aux = aux.append({
                'Ticker': row['CodigodoAtivo'], 
                'Emissor': row['Empresa'], 
                'Infraestrutura': infra, 
                'Data de Saida/Nova Data de Vencimento': row['DatadeSaida/NovoVencimento'],
                'Garantia/Especie': row['Garantia/Especie'], 
                'Valor Nominal na Emissão': row['ValorNominalnaEmissao'],
                'Quantidade no Mercado': row['QuantidadeemMercado'],
                'Índice': ind, 
                'Percentual Multiplicador/Rentabilidade': perct,
                'CNPJ': row['CNPJ'],
                'Deb. Incentivada (Lei12.431)': row['Deb.Incent.(Lei12.431)'],
                'Resgate Antecipado': row['ResgateAntecipado'],
                'Kinea': kinea}, ignore_index=True)

    df = df.append(aux, ignore_index=True) # Junta os dados novos com os antigos

    # Garatindo que todas as do mesmo emissor estão inclusas em Infraestrutura
    aux = df
    for emissor, df_aux in aux.groupby('Emissor'):
        if (df_aux['Infraestrutura'] == 1).any():
            df.loc[aux['Emissor'] == emissor, ['Infraestrutura']] = 1

    clear_sheet(sheet) # Limpa a sheet com os dados antigos
    pandas_to_excel(sheet, df) # Preenche a sheet com os dados antigos e novos

def wis_cadastro_infra() -> None:
    """
        Função para preenchimento das células na planilha Cadastro Infra
    """

    print('Populando planilha Cadastro Infra...')
    
    global CLASS_DEB
    global BDINFRA_path

    sheet = 'Cadastro Infra' # Selecionando a sheet a ser manipulada

    df_cad = pd.read_excel(BDINFRA_path, sheet_name='Cadastro')
    df_cad = df_cad[df_cad['Infraestrutura'] == 1] # Seleciona as rows somente das que forem de infra

    df_rat = pd.read_excel(BDINFRA_path, sheet_name='Rating')

    df = pd.read_excel(BDINFRA_path, sheet_name=sheet)
    df = df[df['Ticker'].notna()] # Retira possíveis vazios

    aux = pd.DataFrame(columns=df.columns)
    for idx_row, row in df_cad.iterrows():

        indexador = row['Índice']
        incentivada = np.nan
        setor = np.nan
        subsetor = np.nan
        ag_rat = np.nan
        rat_emi = np.nan
        
        # Checa se o ticker é Incentivada ou nao
        if row['Índice'] == 'IPCA':
            incentivada = 1
        else:  
            incentivada = 0

        # Checa se o emissor está na planilha Classficação de Debênture para preencher o setor e subsetor
        if row['Emissor'] in list(CLASS_DEB['Emissor']):
            setor = CLASS_DEB[CLASS_DEB['Emissor'] == row['Emissor']]['Setor'].iloc[0]
            subsetor = CLASS_DEB[CLASS_DEB['Emissor'] == row['Emissor']]['Subsetor'].iloc[0]

        # Checa se o ticker está na planilha Rating para preencher Agência e Rating
        if row['Ticker'] in list(df_rat['Ticker']):
            dt_eval = df_rat[df_rat['Ticker'] == row['Ticker']].sort_values('Dia')['Dia'].iloc[0]
            ag_rat = df_rat[(df_rat['Ticker'] == row['Ticker']) & (df_rat['Dia'] == dt_eval)]['Agência de Rating'].iloc[0]
            rat_emi = df_rat[(df_rat['Ticker'] == row['Ticker']) & (df_rat['Dia'] == dt_eval)]['Rating'].iloc[0]

        # Chega para não colocar uma row vazia
        if (indexador != np.nan):
            # Preenchimento das colunas
            aux = aux.append({
                'Ticker': row['Ticker'],
                'Indexador': indexador, 
                'Incentivada': incentivada,
                'Emissor': row['Emissor'],
                'Setor': setor,
                'Subsetor': subsetor,
                'Flag Consolidado': np.nan,
                'Flag Resumo': np.nan,
                'Agência rating': ag_rat,
                'Rating de emissão': rat_emi}, ignore_index=True)

    df = aux # Atualiza os dados novos com os antigos
    clear_sheet(sheet) # Limpa a sheet com os dados antigos
    pandas_to_excel(sheet, df) # Preenche a sheet com os dados antigos e novos

def wis_rating(dt_rat: pd.Timestamp) -> None:
    """
        Função para preenchimento das células na planilha Rating

        :param dt_rat: Data a ser utilizada na sheet de Rating
    """

    print('Populando planilha Rating...')

    global RATING
    global BDINFRA_path

    sheet = 'Rating' # Selecionando a sheet a ser manipulada

    df_cad = pd.read_excel(BDINFRA_path, sheet_name='Cadastro')

    df = pd.read_excel(BDINFRA_path, sheet_name=sheet)
    df = df[df['Ticker'].notna()] # Retira possíveis vazios
    
    aux = pd.DataFrame(columns=df.columns)
    for idx_row, row in df_cad.iterrows():
        # Checa se o ticker não está presente na sheet ou se, estando, está sendo uma data diferente para popular
        if (not (row['Ticker'] in list(df['Ticker']))) or (not (dt_rat in list(df[df['Ticker'] == row['Ticker']]['Dia']))):
            ag_ret = np.nan
            rat = np.nan
            rat_eq = np.nan
            # Checa se o ticker está nas planilha RATING
            if row['Ticker'] in list(RATING['Emissor']):
                ag_ret = RATING[RATING['Emissor'] == row['Ticker']]['Agência'].iloc[0]
                rat = RATING[RATING['Emissor'] == row['Ticker']]['Rating'].iloc[0]
                rat_eq = RATING[RATING['Emissor'] == row['Ticker']]['Rating Equivalente'].iloc[0]
                # Preenchimento das colunas
                aux = aux.append({
                    'Dia': str(dt_rat.date())[-2:] + '/' + str(dt_rat.date())[5:7] + '/' + str(dt_rat.date())[0:4],
                    'Ticker': row['Ticker'], 
                    'Agência de Rating': ag_ret,
                    'Rating': rat,
                    'Rating Equivalente': rat_eq}, ignore_index=True)
				
    df = df.append(aux, ignore_index=True) # Junta os dados novos com os antigos

    clear_sheet(sheet) # Limpa a sheet com os dados antigos
    pandas_to_excel(sheet, df) # Preenche a sheet com os dados antigos e novos

def wis_ipca_anbima(TAXAS_IPCA, IMAB, ETTJ, date) -> None:
    """
        Função para preenchimento das células na planilha IPCA - ANBIMA
    """

    print('Populando planilha IPCA - ANBIMA...')

    global BDINFRA_path

    sheet = 'IPCA - ANBIMA' # Selecionando a sheet a ser manipulada

    ETTJ = ETTJ[ETTJ.index.get_level_values(0) == 'ETTJ IPCA']

    df_cad = pd.read_excel(BDINFRA_path, sheet_name='Cadastro')
    df_cad = df_cad[df_cad['Índice'] == 'IPCA'] # Seleciona as rows somente das que forem com o indexador em IPCA

    df = pd.read_excel(BDINFRA_path, sheet_name=sheet)
    df = df[df['Ticker'].notna()] # Retira possíveis vazios

    aux = pd.DataFrame(columns=df.columns)
    for idx_row, row in df_cad.iterrows(): 
        # Checa se o ticker não está presente na sheet ou se, estando, está sendo uma data diferente para popular
        if (not (row['Ticker'] in list(df['Ticker']))) or (not (date in list(df[df['Ticker'] == row['Ticker']]['Dia']))):
            
            # Checa se o ticker está nas planilha TAXAS_IPCA
            if row['Ticker'] in set(TAXAS_IPCA.index.get_level_values(level=0)):
                ticker = row['Ticker']
                pu_anbima = float(TAXAS_IPCA[TAXAS_IPCA.index.get_level_values(0) == row['Ticker']].loc[(row['Ticker'], 'PU')])
                taxa_anbima = float(TAXAS_IPCA[TAXAS_IPCA.index.get_level_values(0) == row['Ticker']].loc[(row['Ticker'], 'Taxa Indicativa')])/100
                durat_anbima = TAXAS_IPCA[TAXAS_IPCA.index.get_level_values(0) == row['Ticker']].loc[(row['Ticker'], 'Duration')]
                durat_anbima = float(durat_anbima) if ~np.isnan(durat_anbima) else np.nan
                # Caso tenha um valor para duration
                if durat_anbima != np.nan:  
                    # Cálculo do valor mais prócimo de duration disponível
                    durat_ref_ntnb = IMAB[IMAB.index.get_level_values(level=2) == 'Duration'][(np.abs(np.array(IMAB[IMAB.index.get_level_values(level=2) == 'Duration']).astype(float) - durat_anbima)).argmin()] 
                    # Filta a data de vencimento correspondente
                    ntnb_ref_anbima_date = IMAB[(IMAB.index.get_level_values(level=2) == 'Duration') & (IMAB == durat_ref_ntnb)].index.get_level_values(1)[0]
                    # Filta a taxa indicativa correspondente
                    ntnb_ref_anbima_tax = float(IMAB[(IMAB.index.get_level_values(level=2) == 'Expectativa') & (IMAB.index.get_level_values(level=1) == ntnb_ref_anbima_date)].iloc[0])/100
                    # Cálculo do spread entre anbima e ntnb
                    spread_anbima_ntnb = ((1+taxa_anbima)/(1+ntnb_ref_anbima_tax) - 1)
                    # Interpolação entre os durations de ettj e suas taxas ettj ipca e o duration usado
                    ettj_ref_anbima_tax = np.interp(durat_anbima, np.array(ETTJ.index.get_level_values(1)).astype(int), np.array(ETTJ).astype(float))/100
                    # Cálculo do spread entre anbima e ettj
                    spread_anbima_ettj = ((1+taxa_anbima)/(1+ettj_ref_anbima_tax) - 1)
                # Preenchimento das colunas
                aux = aux.append({
                    'Ticker': ticker,
                    'Dia': date, 
                    'PU ANBIMA': pu_anbima,
                    'Duration ANBIMA': durat_anbima,
                    'Taxa ANBIMA': taxa_anbima,
                    'NTN-B Ref ANBIMA': ntnb_ref_anbima_date,
                    'Taxa NTN-B Ref ANBIMA': ntnb_ref_anbima_tax,
                    'Spread B Ref ANBIMA': spread_anbima_ntnb,
                    'Taxa ETTJ Ref ANBIMA': ettj_ref_anbima_tax,
                    'Spread ETTJ Ref ANBIMA': spread_anbima_ettj}, ignore_index=True)

    df = df.append(aux, ignore_index=True) # Junta os dados novos com os antigos

    clear_sheet(sheet) # Limpa a sheet com os dados antigos
    pandas_to_excel(sheet, df) # Preenche a sheet com os dados antigos e novos

def wis_ipca_mercado(REUNE, IMAB, ETTJ, TAXAS_IPCA, date) -> None:
    """
        Função para preenchimento das células na planilha IPCA - Mercado
    """

    print('Populando planilha IPCA - Mercado...')

    global BDINFRA_path

    sheet = 'IPCA - Mercado' # Selecionando a sheet a ser manipulada

    ETTJ = ETTJ[ETTJ.index.get_level_values(0) == 'ETTJ IPCA']
    
    df_cad = pd.read_excel(BDINFRA_path, sheet_name='Cadastro')
    df_cad = df_cad[df_cad['Índice'] == 'IPCA'] # Seleciona as rows somente das que forem com o indexador em IPCA

    df = pd.read_excel(BDINFRA_path, sheet_name=sheet)
    df = df[df['Ticker'].notna()] # Retira possíveis vazios

    aux = pd.DataFrame(columns=df.columns)
    for idx_row, row in df_cad.iterrows():
        # Checa se o ticker não está presente na sheet ou se, estando, está sendo uma data diferente para popular 
        if (not (row['Ticker'] in list(df['Ticker']))) or (not (date in list(df[df['Ticker'] == row['Ticker']]['Dia']))):
            
            # Checa se o ticker está nas planilha REUNE
            if row['Ticker'] in list(REUNE['CETIP']):
                ticker = row['Ticker']
                vol_neg =  REUNE[REUNE['CETIP'] == row['Ticker']]['Faixa de Volume'].iloc[0]
                pu_mercado = float(REUNE[REUNE['CETIP'] == row['Ticker']]['Preço Médio'].iloc[0])
                taxa_mercado = REUNE[REUNE['CETIP'] == row['Ticker']]['Taxa Média'].iloc[0]
                taxa_mercado = float(taxa_mercado)/100 if taxa_mercado != '--' else np.nan
                # Por enquanto, utilizando o duration da ANBIMA
                durat_mercado = TAXAS_IPCA[TAXAS_IPCA.index.get_level_values(0) == row['Ticker']].loc[(row['Ticker'], 'Duration')] if row['Ticker'] in list(TAXAS_IPCA.index.get_level_values(0)) else np.nan
                durat_mercado = float(durat_mercado) if ~np.isnan(durat_mercado) else np.nan  
                # Caso tenha um valor para duration e de taxa indicativa
                if (durat_mercado != np.nan) and (taxa_mercado != np.nan):
                    # Cálculo do valor mais prócimo de duration disponível
                    durat_ref_ntnb = IMAB[IMAB.index.get_level_values(level=2) == 'Duration'][(np.abs(np.array(IMAB[IMAB.index.get_level_values(level=2) == 'Duration']).astype(float) - durat_mercado)).argmin()]
                    # Filta a data de vencimento correspondente
                    ntnb_ref_mercado_date = IMAB[(IMAB.index.get_level_values(level=2) == 'Duration') & (IMAB == durat_ref_ntnb)].index.get_level_values(1)[0]
                    # Filta a taxa indicativa correspondente
                    ntnb_ref_mercado_tax = float(IMAB[(IMAB.index.get_level_values(level=2) == 'Expectativa') & (IMAB.index.get_level_values(level=1) == ntnb_ref_mercado_date)].iloc[0])/100
                    # Cálculo do spread entre mercado e ntnb
                    spread_mercado_ntnb = ((1+taxa_mercado)/(1+ntnb_ref_mercado_tax) - 1)
                    # Interpolação entre os durations de ettj e suas taxas ettj ipca e o duration usado
                    ettj_ref_mercado_tax = np.interp(durat_mercado, np.array(ETTJ.index.get_level_values(1)).astype(int), np.array(ETTJ).astype(float))/100
                    # Cálculo do spread entre mercado e ettj
                    spread_mercado_ettj = ((1+taxa_mercado)/(1+ettj_ref_mercado_tax) - 1)
                # Preenchimento das colunas
                aux = aux.append({
                    'Ticker': ticker,
                    'Dia': date, 
                    'Volume Negociado': vol_neg,
                    'PU Mercado': pu_mercado,
                    'Duration Mercado': durat_mercado,
                    'Taxa Mercado': taxa_mercado,
                    'NTN-B Ref Mercado': ntnb_ref_mercado_date,
                    'Taxa NTN-B Ref Mercado': ntnb_ref_mercado_tax,
                    'Spread B Ref Mercado': spread_mercado_ntnb,
                    'Taxa ETTJ Ref Mercado': ettj_ref_mercado_tax,
                    'Spread ETTJ Ref Mercado ': spread_mercado_ettj}, ignore_index=True)

    df = df.append(aux, ignore_index=True) # Junta os dados novos com os antigos

    clear_sheet(sheet) # Limpa a sheet com os dados antigos
    pandas_to_excel(sheet, df) # Preenche a sheet com os dados antigos e novos

def wis_cdi_anbima(TAXAS_CDI, date) -> None:
    """
        Função para preenchimento das células na planilha CDI - ANBIMA
    """

    print('Populando planilha CDI - ANBIMA...')

    global BDINFRA_path
    
    sheet = 'CDI - ANBIMA' # Selecionando a sheet a ser manipulada

    df_cad = pd.read_excel(BDINFRA_path, sheet_name='Cadastro')
    df_cad = df_cad[df_cad['Índice'] == 'CDI'] # Seleciona as rows somente das que forem com o indexador em CDI

    df = pd.read_excel(BDINFRA_path, sheet_name=sheet)
    df = df[df['Ticker'].notna()] # Retira possíveis vazios

    aux = pd.DataFrame(columns=df.columns)
    for idx_row, row in df_cad.iterrows(): 
        # Checa se o ticker não está presente na sheet ou se, estando, está sendo uma data diferente para popular
        if (not (row['Ticker'] in list(df['Ticker']))) or (not (date in list(df[df['Ticker'] == row['Ticker']]['Dia']))):

            # Checa se o ticker está nas planilha TAXAS_CDI
            if row['Ticker'] in set(TAXAS_CDI.index.get_level_values(level=0)):
                ticker = row['Ticker']
                pu_anbima = TAXAS_CDI[TAXAS_CDI.index.get_level_values(0) == row['Ticker']].loc[(row['Ticker'], 'PU')]
                pu_anbima = float(pu_anbima) if ~np.isnan(pu_anbima) else np.nan
                durat_anbima = TAXAS_CDI[TAXAS_CDI.index.get_level_values(0) == row['Ticker']].loc[(row['Ticker'], 'Duration')]
                durat_anbima = float(durat_anbima) if ~np.isnan(durat_anbima) else np.nan
                taxa_anbima = TAXAS_CDI[TAXAS_CDI.index.get_level_values(0) == row['Ticker']].loc[(row['Ticker'], 'Taxa Indicativa')]
                taxa_anbima = float(taxa_anbima)/100 if ~np.isnan(taxa_anbima) else np.nan
                spread_anbima = taxa_anbima
                # Preenchimento das colunas
                aux = aux.append({
                    'Ticker': ticker,
                    'Dia': date, 
                    'PU ANBIMA': pu_anbima,
                    'Duration ANBIMA': durat_anbima,
                    'Taxa ANBIMA': taxa_anbima,
                    'Spread ANBIMA': spread_anbima}, ignore_index=True)

    df = df.append(aux, ignore_index=True) # Junta os dados novos com os antigos

    clear_sheet(sheet) # Limpa a sheet com os dados antigos
    pandas_to_excel(sheet, df) # Preenche a sheet com os dados antigos e novos

def wis_cdi_mercado(REUNE, IMAB, date) -> None:
    """
        Função para preenchimento das células na planilha CDI - Mercado
    """

    print('Populando planilha CDI - Mercado...')

    global BDINFRA_path

    sheet = 'CDI - Mercado' # Selecionando a sheet a ser manipulada

    df_cad = pd.read_excel(BDINFRA_path, sheet_name='Cadastro')
    df_cad = df_cad[df_cad['Índice'] == 'CDI'] # Seleciona as rows somente das que forem com o indexador em CDI

    df = pd.read_excel(BDINFRA_path, sheet_name=sheet)
    df = df[df['Ticker'].notna()] # Retira possíveis vazios

    aux = pd.DataFrame(columns=df.columns)
    for idx_row, row in df_cad.iterrows(): 
        # Checa se o ticker não está presente na sheet ou se, estando, está sendo uma data diferente para popular
        if (not (row['Ticker'] in list(df['Ticker']))) or (not (date in list(df[df['Ticker'] == row['Ticker']]['Dia']))):

            # Checa se o ticker está nas planilha REUNE
            if row['Ticker'] in list(REUNE['CETIP']):
                ticker = row['Ticker']
                vol_neg = REUNE[REUNE['CETIP'] == row['Ticker']]['Faixa de Volume'].iloc[0]
                pu_mercado = float(REUNE[REUNE['CETIP'] == row['Ticker']]['Preço Médio'].iloc[0])
                taxa_mercado = REUNE[REUNE['CETIP'] == row['Ticker']]['Taxa Média'].iloc[0]
                taxa_mercado = float(taxa_mercado) if taxa_mercado != '--' else np.nan
                spread_mercado = taxa_mercado
                # Caso tenha um valor de taxa indicativa
                if (taxa_mercado != np.nan):
                    # Interpolação entre as taxas de ntnb e os durations e a taxa usada
                    durat_mercado = np.interp(taxa_mercado, np.array(IMAB[IMAB.index.get_level_values(2) == 'Expectativa']).astype(float), np.array(IMAB[IMAB.index.get_level_values(2) == 'Duration']).astype(int))
                # Preenchimento das colunas
                aux = aux.append({
                    'Ticker': ticker,
                    'Dia': date, 
                    'Volume Negociado': vol_neg,
                    'PU Mercado': pu_mercado,
                    'Duration Mercado': durat_mercado,
                    'Taxa Mercado': taxa_mercado,
                    'Spread Mercado': spread_mercado}, ignore_index=True)

    df = df.append(aux, ignore_index=True) # Junta os dados novos com os antigos

    clear_sheet(sheet) # Limpa a sheet com os dados antigos
    pandas_to_excel(sheet, df) # Preenche a sheet com os dados antigos e novos

def wis_pct_cdi_anbima(TAXAS_PCT_CDI, Pre_REF, ETTJ, date) -> None: 
    """
        Função para preenchimento das células na planilha %CDI - ANBIMA
    """

    print('Populando planilha %CDI - ANBIMA...')

    global BDINFRA_path

    sheet = '%CDI - ANBIMA' # Selecionando a sheet a ser manipulada

    ETTJ = ETTJ[ETTJ.index.get_level_values(0) == 'ETTJ PRE']

    Pre_REF_dict = {'F': '01', 'G': '02', 'H': '03', 'J': '04', 'K': '05', 'M': '06', 'N': '07', 'Q': '08', 'U': '09', 'V': '10', 'X': '11', 'Z': '12'} # Conversão de letras e números

    df_cad = pd.read_excel(BDINFRA_path, sheet_name='Cadastro')
    df_cad = df_cad[df_cad['Índice'] == '%CDI'] # Seleciona as rows somente das que forem com o indexador em %CDI

    df = pd.read_excel(BDINFRA_path, sheet_name=sheet)
    df = df[df['Ticker'].notna()] # Retira possíveis vazios

    aux = pd.DataFrame(columns=df.columns)
    for idx_row, row in df_cad.iterrows(): 
        # Checa se o ticker não está presente na sheet ou se, estando, está sendo uma data diferente para popular
        if (not (row['Ticker'] in list(df['Ticker']))) or (not (date in list(df[df['Ticker'] == row['Ticker']]['Dia']))):

            # Checa se o ticker está nas planilha TAXAS_PCT_CDI
            if row['Ticker'] in set(TAXAS_PCT_CDI.index.get_level_values(level=0)):
                ticker = row['Ticker']
                taxa_emi = row['Percentual Multiplicador/Rentabilidade']
                pu_anbima = TAXAS_PCT_CDI[TAXAS_PCT_CDI.index.get_level_values(0) == row['Ticker']].loc[(row['Ticker'], 'PU')]
                pu_anbima = float(pu_anbima) if ~np.isnan(pu_anbima) else np.nan
                durat_anbima = TAXAS_PCT_CDI[TAXAS_PCT_CDI.index.get_level_values(0) == row['Ticker']].loc[(row['Ticker'], 'Duration')]
                durat_anbima = float(durat_anbima) if ~np.isnan(durat_anbima) else np.nan
                taxa_anbima = TAXAS_PCT_CDI[TAXAS_PCT_CDI.index.get_level_values(0) == row['Ticker']].loc[(row['Ticker'], 'Taxa Indicativa')]
                taxa_anbima = float(taxa_anbima)/100 if ~np.isnan(taxa_anbima) else np.nan
                # Caso tenha um valor de duration
                if durat_anbima != np.nan:
                    # Busca da data com intervalo em BDs mais próximo
                    date_from = date[-4:] + '-' + date[3:5] + '-' + date[:2]
                    dates_to = list(Pre_REF[Pre_REF['Data'] == date_from].dropna(axis=1).columns[1:])
                    pre_ref_list = [] # Guarda o intervalo de BDs para todas as datas 
                    for dt in dates_to:
                        date_to = '20' + dt[-2:] + '-' + Pre_REF_dict[dt[-3:-2]] + '-01'
                        pre_ref_list.append(np.busday_count(date_from, date_to, holidays = Feriados))
                    pre_ref_col = np.array(dates_to)[(np.abs(np.array(pre_ref_list).astype(float) - durat_anbima)).argmin()] # Descobre data
                    pre_ref_anbima_date = Pre_REF[Pre_REF['Data'] == date_from][pre_ref_col].iloc[0] # Taxa para determinado dia
                    pre_ref_anbima_tax = taxa_anbima-1
                    spread_anbima_pre = pre_ref_anbima_tax*pre_ref_anbima_date
                    # Interpolação entre os durations de ettj e seus taxas ettj pref e o duration usado
                    ettj_ref_anbima_tax = np.interp(durat_anbima, np.array(ETTJ.index.get_level_values(1)).astype(int), np.array(ETTJ).astype(float))/100
                    # Cálculo do spread entre anbima e ettj
                    spread_anbima_ettj = ((1+taxa_anbima)/(1+ettj_ref_anbima_tax) - 1)
                # Preenchimento das colunas
                aux = aux.append({
                    'Ticker': ticker,
                    'Dia': date, 
                    'Taxa Emissão': taxa_emi,
                    'PU ANBIMA': pu_anbima,
                    'Duration ANBIMA': durat_anbima,
                    'Taxa ANBIMA': taxa_anbima,
                    'Pré Ref ANBIMA': pre_ref_anbima_date,
                    'Taxa Pré Ref ANBIMA': pre_ref_anbima_tax,
                    'Spread Pré Ref ANBIMA': spread_anbima_pre,
                    'Taxa ETTJ Ref ANBIMA': ettj_ref_anbima_tax,
                    'Spread ETTJ Ref ANBIMA': spread_anbima_ettj}, ignore_index=True)

    df = df.append(aux, ignore_index=True) # Junta os dados novos com os antigos

    clear_sheet(sheet) # Limpa a sheet com os dados antigos
    pandas_to_excel(sheet, df) # Preenche a sheet com os dados antigos e novos

def wis_pct_cdi_mercado(REUNE, TAXAS_PCT_CDI, Pre_REF, ETTJ, date) -> None:
    """
        Função para preenchimento das células na planilha %CDI - Mercado
    """

    print('Populando planilha %CDI - Mercado...')

    global BDINFRA_path

    sheet = '%CDI - Mercado' # Selecionando a sheet a ser manipulada

    Pre_REF_dict = {'F': '01', 'G': '02', 'H': '03', 'J': '04', 'K': '05', 'M': '06', 'N': '07', 'Q': '08', 'U': '09', 'V': '10', 'X': '11', 'Z': '12'} # Conversão de letras e números

    ETTJ = ETTJ[ETTJ.index.get_level_values(0) == 'ETTJ PRE']

    df_cad = pd.read_excel(BDINFRA_path, sheet_name='Cadastro')
    df_cad = df_cad[df_cad['Índice'] == '%CDI'] # Seleciona as rows somente das que forem com o indexador em %CDI

    df = pd.read_excel(BDINFRA_path, sheet_name=sheet)
    df = df[df['Ticker'].notna()] # Retira possíveis vazios

    aux = pd.DataFrame(columns=df.columns)
    for idx_row, row in df_cad.iterrows(): 
        # Checa se o ticker não está presente na sheet ou se, estando, está sendo uma data diferente para popular
        if (not (row['Ticker'] in list(df['Ticker']))) or (not (date in list(df[df['Ticker'] == row['Ticker']]['Dia']))):

            # Checa se o ticker está nas planilha REUNE
            if row['Ticker'] in list(REUNE['CETIP']):
                ticker = row['Ticker']
                taxa_emi = row['Percentual Multiplicador/Rentabilidade']
                vol_neg = REUNE[REUNE['CETIP'] == row['Ticker']]['Faixa de Volume'].iloc[0]
                pu_mercado = float(REUNE[REUNE['CETIP'] == row['Ticker']]['Preço Médio'].iloc[0])
                durat_mercado = TAXAS_PCT_CDI[TAXAS_PCT_CDI.index.get_level_values(0) == row['Ticker']].loc[(row['Ticker'], 'Duration')] if row['Ticker'] in list(TAXAS_PCT_CDI.index.get_level_values(0)) else np.nan
                durat_mercado = float(durat_mercado) if ~np.isnan(durat_mercado) else np.nan
                taxa_mercado = REUNE[REUNE['CETIP'] == row['Ticker']]['Taxa Média'].iloc[0]
                taxa_mercado = float(taxa_mercado)/100 if taxa_mercado != '--' else np.nan
                # Caso tenha um valor para duration e de taxa indicativa
                if (durat_mercado != np.nan) and (taxa_mercado != np.nan):
                    # Busca da data com intervalo em BDs mais próximo
                    date_from = date[-4:] + '-' + date[3:5] + '-' + date[:2]
                    dates_to = list(Pre_REF[Pre_REF['Data'] == date_from].dropna(axis=1).columns[1:])
                    pre_ref_list = [] # Guarda o intervalo de BDs para todas as datas 
                    for dt in dates_to:
                        date_to = '20' + dt[-2:] + '-' + Pre_REF_dict[dt[-3:-2]] + '-01'
                        pre_ref_list.append(np.busday_count(date_from, date_to, holidays = Feriados))
                    pre_ref_col = np.array(dates_to)[(np.abs(np.array(pre_ref_list).astype(float) - durat_mercado)).argmin()] # Descobre data
                    pre_ref_mercado_date = Pre_REF[Pre_REF['Data'] == date_from][pre_ref_col].iloc[0] # Taxa para determinado dia
                    pre_ref_mercado_tax = taxa_mercado-1
                    spread_mercado_pre = pre_ref_mercado_tax*pre_ref_mercado_date
                    # Interpolação entre os durations de ettj e seus taxas ettj pref e o duration usado
                    ettj_ref_mercado_tax = np.interp(durat_mercado, np.array(ETTJ.index.get_level_values(1)).astype(int), np.array(ETTJ).astype(float))/100
                    # Cálculo do spread entre mercado e ettj
                    spread_mercado_ettj = ((1+taxa_mercado)/(1+ettj_ref_mercado_tax) - 1)
                # Preenchimento das colunas
                aux = aux.append({
                    'Ticker': ticker,
                    'Dia': date, 
                    'Taxa Emissão': taxa_emi,
                    'Volume Negociado': vol_neg, 
                    'PU Mercado': pu_mercado,
                    'Duration Mercado': durat_mercado,
                    'Taxa Mercado': taxa_mercado,
                    'Pré Ref Mercado': pre_ref_mercado_date,
                    'Taxa Pré Ref Mercado': pre_ref_mercado_tax,
                    'Spread Pré Ref Mercado': spread_mercado_pre,
                    'Taxa ETTJ Ref Mercado': ettj_ref_mercado_tax,
                    'Spread ETTJ Ref Mercado': spread_mercado_ettj}, ignore_index=True)

    df = df.append(aux, ignore_index=True) # Junta os dados novos com os antigos

    clear_sheet(sheet) # Limpa a sheet com os dados antigos
    pandas_to_excel(sheet, df) # Preenche a sheet com os dados antigos e novos

source()
wis_cadastro()
wis_cadastro_infra()
wis_rating(dt_rat)

dates = [file for file in os.listdir(downloads_path) if file.isnumeric()]

for dt in dates:
    print('Inserindo para data {}'.format(dt))

    date = dt[:2] + '/' + dt[2:4] + '/' + dt[-4:]
    semester = '1' if pd.to_datetime(date, dayfirst = True).quarter in [1,2] else '2'

    TAXAS = pd.read_excel(os.path.join(downloads_path,  date_to_file(date, 'TAXAS', semester)))
    TAXAS, Lixo = clean(TAXAS, 'TAXAS', date)

    IMAB = pd.read_excel(os.path.join(downloads_path,  date_to_file(date, 'IMAB', semester)))
    IMAB, Lixo = clean(IMAB, 'IMAB', date)

    Pre_REF = pd.read_excel(os.path.join(downloads_path,  date_to_file(date, 'Pre_REF', semester)))
    Pre_REF, Lixo = clean(Pre_REF, 'Pre_REF', date)

    ETTJ = pd.read_excel(os.path.join(downloads_path,  date_to_file(date, 'ETTJ', semester)))
    ETTJ, Lixo = clean(ETTJ, 'ETTJ', date)

    REUNE = pd.read_csv(os.path.join(downloads_path, dt, date_to_file(date, 'REUNE', semester)), sep=";",header=2, encoding='ANSI')
    REUNE, Lixo = clean(REUNE, 'REUNE', date)

    wis_ipca_anbima(TAXAS, IMAB, ETTJ, date)
    wis_ipca_mercado(REUNE, IMAB, ETTJ, TAXAS, date)
    wis_cdi_anbima(TAXAS, date)
    wis_cdi_mercado(REUNE, IMAB, date)
    wis_pct_cdi_anbima(TAXAS, Pre_REF, ETTJ, date)
    wis_pct_cdi_mercado(REUNE, TAXAS, Pre_REF, ETTJ, date)